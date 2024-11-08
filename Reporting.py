import cx_Oracle
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def modify_cells_in_workbook(file_path, sheet_name, cells_to_modify):
    workbook = openpyxl.load_workbook(file_path)
    if sheet_name not in workbook.sheetnames:
        print(f"Sheet {sheet_name} not found in {file_path}.")
        return
    sheet = workbook[sheet_name]
    for cell_address, new_value in cells_to_modify.items():
        cell = sheet[cell_address]
        if cell.coordinate in sheet.merged_cells:
            merged_range = [rng for rng in sheet.merged_cells.ranges if cell.coordinate in rng][0]
            sheet.unmerge_cells(merged_range.coord)
            sheet[cell_address] = new_value
            sheet.merge_cells(merged_range.coord)
        else:
            sheet[cell_address] = new_value
    output_file_name = 'output_file_data.xlsx'  # Change to desired output file name
    workbook.save(output_file_name)


def read_db_config(config_file_path):
    config = {}
    with open(config_file_path, 'r') as file:
        for line in file:
            key, value = line.strip().split('=')
            config[key] = value
    return config


def create_oracle_connection(username, password, host, port, service_name):
    dsn = cx_Oracle.makedsn(host, port, service_name=service_name)
    connection = cx_Oracle.connect(username, password, dsn)
    return connection


def execute_sql_query(connection, sql_query, date_param):
    cursor = connection.cursor()
    cursor.execute(sql_query, {'date_param': date_param})
    result = cursor.fetchall()
    cursor.close()
    return result


def close_connection(connection):
    connection.close()


def read_sql_query(sheet_name, cell_ref):
    file_path = os.path.join('SQL_Scripts', sheet_name, f'{cell_ref}.txt')
    if not os.path.exists(file_path):
        print(f"SQL file for cell {cell_ref} not found in sheet {sheet_name}.")
        return None
    with open(file_path, 'r') as file:
        return file.read()


def execute_sql_for_all_sheets(connection, date_param):
    # Get list of sheet subfolders in 'SQL_Scripts'
    folder_path = 'SQL_Scripts'
    sheet_folders = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]

    for sheet_name in sheet_folders:
        # Get list of SQL files in the current sheet's subfolder
        sheet_folder_path = os.path.join(folder_path, sheet_name)
        sql_files = [f for f in os.listdir(sheet_folder_path) if f.endswith('.txt')]

        cells_to_modify = {}

        for sql_file in sql_files:
            cell_ref = sql_file.replace('.txt', '')  # Extract cell reference from file name
            sql_query = read_sql_query(sheet_name, cell_ref)
            if sql_query:
                results = execute_sql_query(connection, sql_query, date_param)
                if results:
                    print(f"Result for {cell_ref} in sheet {sheet_name}: {results[0][0]}")
                    cells_to_modify[cell_ref] = results[0][0]
                else:
                    print(f"No result found for {cell_ref} in sheet {sheet_name}")

        file_path = 'data_etr2.xlsx'  # Change to your file path
        modify_cells_in_workbook(file_path, sheet_name, cells_to_modify)

    close_connection(connection)


if __name__ == "__main__":
    config = read_db_config('db_config.txt')
    connection = create_oracle_connection(
        config['username'],
        config['password'],
        config['host'],
        config['port'],
        config['service_name']
    )
    date_param = "01/23/2023"  # Modify this line to set the desired date parameter
    execute_sql_for_all_sheets(connection, date_param)
